# -*- coding: utf-8 -*-
# 封装excel
# @Time    : 2019/6/26 20:44
# @Author  : YoYo

from openpyxl import load_workbook, Workbook


class HandleExcel:
    """Excel操作封装类"""

    def __init__(self):
        self.filename = None
        self.workbook = None
        self.worksheet = None
        self.sheet_name = None

    def get_excel(self, filename):
        self.filename = filename
        self.workbook = load_workbook(self.filename)
        if self.workbook is None:
            self.workbook = Workbook()
        else:
            self.workbook = load_workbook(self.filename)

    def get_exist_excel(self, filename):
        """
        获取已经存在的excel
        :param filename: 文件名
        :return:
        """
        self.filename = filename
        self.workbook = load_workbook(self.filename)

    def create_excel(self, filename):
        """
        创建一个Excel文件
        :param filename: 文件名
        :return:
        """
        try:
            self.filename = filename
            self.workbook = Workbook()
        except  FileNotFoundError as e:
            print("报错的信息是：", e)
            raise e

    # def get_cases(self, sheet_name):
    #     """获取用例数据"""
    #     sheet = self.workbook[sheet_name]
    #     max_row = sheet.max_row  # 获取所在sheet页的最大行
    #     cases = []
    #     for i in range(2, max_row + 1):
    #         case = Case()  # 实例化一个case对象，用来存放测试数据
    #         case.case_id = sheet.cell(row=i, column=1).value
    #         case.title = sheet.cell(row=i, column=2).value
    #         case.method = sheet.cell(row=i, column=3).value
    #         case.url = sheet.cell(row=i, column=4).value
    #         case.param = sheet.cell(row=i, column=5).value
    #         case.expected = sheet.cell(row=i, column=6).value
    #         cases.append(case)
    #     return cases

    def get_cases(self, sheet_name):
        """
        获取所有测试用例数据并以字典组成的列表样式返回数据
        :return:
        """
        self.sheet_name = sheet_name
        if self.sheet_name is None:
            self.worksheet = self.workbook.active
        else:
            self.worksheet = self.workbook[self.sheet_name]
        my_list = []
        first_data = tuple(self.worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        titles = first_data[0]
        for data in tuple(self.worksheet.iter_rows(min_row=2, values_only=True)):
            my_list.append(dict(zip(titles, data)))
        return my_list

    def get_sheet_names(self):
        """
        获取到workbook里面所有sheet名称的列表
        :return:
        """
        return self.workbook.sheetnames

    def write_text_by_case_id(self, sheet_name, case_id, actual, result):
        """
        根据sheet name定位到sheet，然后根据case id 定位到行，取到当前行里面的actual这个单元格,然后给它赋值，最后保存当前workbook
        :param sheet_name:
        :param case_id:
        :param actual:
        :param result:
        :return:
        """
        sheet = self.workbook[sheet_name]
        max_row = sheet.max_row
        for i in range(2, max_row + 1):
            case_id_i = sheet.cell(i, 1).value  # 获取第 i 行第一列，也就是case_id这一列
            if case_id_i == case_id:  # 判断取到Excel里面渠道的当前行的case_id 是否等于传进来的case_id
                sheet.cell(i, 7).value = actual  # 写入传进来的actual到当前的actual列的单元格
                sheet.cell(i, 8).value = result  # 写入传进来的actual到当前的result列的单元格
                self.workbook.save(self.filename)
                break

    # def save_data(self, row, column, data=None):
    #     """
    #     保存数据
    #     :param row: 行号
    #     :param column: 列号
    #     :param data: 单元格内值
    #     """
    #     wb = load_workbook(self.file_name)
    #     if self.sheet_name is None:
    #         ws = wb.active
    #     else:
    #         ws = wb[self.sheet_name]
    #     ws.cell(row, column, data)
    #     wb.save(self.file_name)
    #     wb.close()
    #
    def save_list_data(self, sheet_name, data_list):
        """
        保存以字典组成的列表数据到excel中
       :param data_list: 列表数据
        """
        if data_list is None:
            return
        self.sheet_name = sheet_name
        if self.sheet_name is None:
            self.workbook = self.workbook.active
        else:
            self.worksheet = self.workbook[self.sheet_name]
        titles = data_list[0]
        for i, key in enumerate(titles.keys()):
            self.worksheet.cell(1, i + 1, str(key))
        for j in range(0, len(data_list)):
            data_item = data_list[j]
            if isinstance(data_item, dict):
                for k, data in enumerate(data_item.values()):
                    self.worksheet.cell(j + 2, k + 1, str(data))
        self.workbook.save(self.filename)
        self.workbook.close()
